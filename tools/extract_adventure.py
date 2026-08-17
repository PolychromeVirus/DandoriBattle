#!/usr/bin/env python3
"""Extract the Adventure (Campaign) boards from PikminCardReference.xlsx -> adventure.json.

The 'Campaign Maps' sheet encodes each board as a 5-column x 7-row grid of CELL FILL COLORS
(the row-2 legend maps colour -> space type), with text overlays for structures, enemy/boss
names, and the starting pikmin kit. Boards are grouped into Adventure Scenarios (chapters).

Grid geometry (validated against boards.json via the 'Board Layouts' sheet, same colour scheme):
  - A board label 'ADVx.y' sits at (label_row, label_col); the '(players)' note is one col left.
  - The 5 lane columns are label_col-1 .. label_col+3.
  - The 7 space rows are label_row+1, +3, +5, ... +13 (every other row); pikmin kit at label_row+15.
  - The treasure row is label_row+1 (top / FAR end). Home is at the bottom, so we read the grid
    BOTTOM-UP: idx 0 = nearest home (bottom row), idx 6 = far/treasure (top row). This makes the
    board HOME-ANCHORED (one home, treasure at the far end) - exactly the adventure layout.

Output: datafiles/data/adventure.json  { scenarios: [ { name, timed, boards: [ {id,label,players,
  homeAnchored:true, basicColors, kit, lanes:[[{kind,hazard?}...]x7]x5, structures:[], enemies:[],
  unmapped:[] } ] } ] }
"""
import json, re, os, openpyxl
from openpyxl.utils import get_column_letter

XLSX = os.path.join(os.path.dirname(__file__), '..', 'PikminCardReference.xlsx')
OUT  = os.path.join(os.path.dirname(__file__), '..', 'datafiles', 'data', 'adventure.json')

# colour (ARGB) -> space kind. Both sheets' shades are covered.
COLOR_KIND = {
    'FFB6D7A8': ('plain', None),  'FFB7B7B7': ('enemy', None),
    'FFFF0000': ('hazard', 'fire'), 'FF0000FF': ('hazard', 'water'), 'FF3D85C6': ('hazard', 'water'),
    'FFFCE5CD': ('hazard', 'height'), 'FF00FFFF': ('hazard', 'ice'),
    'FF9900FF': ('hazard', 'poison'), 'FF8E7CC3': ('hazard', 'poison'),
    'FFB45F06': ('hazard', 'chasm'), 'FFFFFF00': ('treasure', None),
}
# overlay structure text -> game structure id (hazards.json). NET TRAP has no game analog yet.
STRUCT = {
    'WALL': 'wall', 'BRIDGE': 'bridge', 'CLIMBING STICK': 'climbingstick', 'CLIMB. STICK': 'climbingstick',
    'TUNNEL': 'tunnel', 'CRYSTAL WALL': 'crystalwall', 'ELECTRIC WALL': 'electricwall',
    'ICE WALL': 'icewall', 'WEBBED WALL': 'arachnodeweb', 'FIRE GEYSER': 'firegeyser',
    'WATER SPOUT': 'waterspout', 'WATER GEYSER': 'waterspout', 'ICE VENT': 'icevent',
    'POISON EMITTER': 'poisonemitter', 'ELECTRICITY GENERATOR': 'electricitygenerator',
    'ELECRICITY GENERATOR': 'electricitygenerator', 'ELECTRICITY GEN.': 'electricitygenerator',
}
PIKMIN = {'RED','BLUE','YELLOW','PURPLE','WHITE','ROCK','WINGED','ICE'}

def load_enemy_ids():
    d = json.load(open(os.path.join(os.path.dirname(__file__), '..', 'datafiles', 'data', 'enemies.json')))
    # map a normalized (letters-only) key -> the real id, so hyphenated ids (bug-eyedcrawmad,
    # man-at-legs) still match a spaced/hyphenated overlay name.
    return {re.sub(r'[^a-z]', '', e['id'].lower()): e['id'] for e in d['enemies']}
ENEMY_IDS = load_enemy_ids()
# typo fixups for names misspelled in the xlsx overlays (normalized letters-only key -> real id)
ENEMY_IDS.setdefault('soverignbulblax', 'sovereignbulblax')   # ADV1.6 boss: "SOVERIGN BULBLAX" (missing E)

def enemy_id(text):
    """UPPER SPACED name -> the game's enemy id (matches enemies.json for most). Returns id or None."""
    return ENEMY_IDS.get(re.sub(r'[^a-z]', '', text.lower()))

DATA = os.path.join(os.path.dirname(__file__), '..', 'datafiles', 'data')
def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())
def _load_list(path, *keys):
    d = json.load(open(path))
    if isinstance(d, list):
        return d
    for k in keys:
        if k in d and isinstance(d[k], list):
            return d[k]
    for v in d.values():           # fall back to the first list value
        if isinstance(v, list):
            return v
    return []

# name -> id maps for the campaign decks (normalized, punctuation/spacing-insensitive)
TREASURE_IDS = {_norm(t['name']): t['id'] for t in _load_list(os.path.join(DATA, 'treasures.json'), 'treasures')}
GATHER_IDS   = {_norm(g['name']): g['id'] for g in _load_list(os.path.join(DATA, 'gather.json'), 'gathers', 'cards')}

MAPKEY = re.compile(r'^A\d\.\d$')   # e.g. 'A1.1' (a board label ADVx.y -> Ax.y)

def parse_enemy_decks(wbv):
    """{ 'A1.1': [{id,count}, ...] } - per-map enemy deck, from the two campaign enemy sheets
    (col H = name, header row picks out the Ax.y map columns; each cell = copies of that enemy)."""
    decks = {}
    for sheet in ('Campaign 1 Enemies', 'Campaign Enemy Cards'):
        if sheet not in wbv.sheetnames:
            continue
        ws = wbv[sheet]
        mapcols = {c: ws.cell(1, c).value.strip() for c in range(1, ws.max_column + 1)
                   if isinstance(ws.cell(1, c).value, str) and MAPKEY.match(ws.cell(1, c).value.strip())}
        for r in range(2, ws.max_row + 1):
            nm = ws.cell(r, 8).value
            if not isinstance(nm, str) or not nm.strip():
                continue
            eid = enemy_id(nm)
            if not eid:
                continue
            for c, mk in mapcols.items():
                cnt = ws.cell(r, c).value
                if isinstance(cnt, (int, float)) and cnt > 0:
                    decks.setdefault(mk, []).append({'id': eid, 'count': int(cnt)})
    return decks

def parse_cull_rules(wbv):
    """{ 'A1.2': {'remove':5,'thr':3,'type':'hp'}, ... } - the between-map ENEMY CULL rules.
    In the campaign enemy sheets, cols R/S/T/U = before / Remove / </= / type. 'before'=1.2 means
    "right before map A1.2 the player removes <Remove> deck cards whose <type> stat is </= <thr>",
    after which that map's own column adds its advanced cards. Keyed by the map the cull precedes."""
    rules = {}
    for sheet in ('Campaign 1 Enemies', 'Campaign Enemy Cards'):
        if sheet not in wbv.sheetnames:
            continue
        ws = wbv[sheet]
        # find the before/Remove/</=/type header columns (don't hard-code R-U for other sheets)
        hdr = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if isinstance(h, str):
                k = h.strip().lower()
                if k in ('before', 'remove', '</=', 'type'):
                    hdr[k] = c
        if not all(k in hdr for k in ('before', 'remove', '</=', 'type')):
            continue
        for r in range(2, ws.max_row + 1):
            bf = ws.cell(r, hdr['before']).value
            rm = ws.cell(r, hdr['remove']).value
            th = ws.cell(r, hdr['</=']).value
            ty = ws.cell(r, hdr['type']).value
            if not isinstance(bf, (int, float)) or not isinstance(rm, (int, float)):
                continue
            mk = 'A' + ('%g' % bf)                               # 1.2 -> 'A1.2'
            rules[mk] = {'remove': int(rm), 'thr': int(th) if isinstance(th, (int, float)) else 0,
                         'type': str(ty).strip().lower() if isinstance(ty, str) else 'hp'}
    return rules

def parse_treasures(wv):
    """{ 'A1.1': [{id,weight,name}, ...] } - the map's treasure pool (Campaign Maps cols Z/AA/AB/AC)."""
    tres = {}
    for r in range(1, wv.max_row + 1):
        nm, mk, wt = wv.cell(r, 26).value, wv.cell(r, 27).value, wv.cell(r, 29).value   # Z, AA, AC
        if isinstance(nm, str) and isinstance(mk, str) and MAPKEY.match(mk.strip()):
            tres.setdefault(mk.strip(), []).append({
                'id': TREASURE_IDS.get(_norm(nm)), 'name': nm.strip(),
                'weight': int(wt) if isinstance(wt, (int, float)) else None})
    return tres

def parse_gather_deck(wbv):
    """The campaign gather deck (uniform across maps): 'Campaign Gather Cards' col A = COUNT of each card,
    col B = name. Expanded to the real ratios (e.g. Raw Material x12) so pairs (Raw Material) actually turn up."""
    ws = wbv['Campaign Gather Cards']
    out = []
    for r in range(1, ws.max_row + 1):
        cnt, nm = ws.cell(r, 1).value, ws.cell(r, 2).value
        if isinstance(nm, str) and nm.strip() and nm.strip().lower() != 'none':
            gid = GATHER_IDS.get(_norm(nm))
            if gid:
                out.extend([gid] * (int(cnt) if isinstance(cnt, (int, float)) and cnt > 0 else 1))
    return out

def parse_event_decks(wbv):
    """{ mapPosition(int) -> [{name,good,n,desc}] } - the random-event deck by board POSITION in a
    scenario. 'Campaign Event Cards' cols D..I = Map 1 .. Map 6 & 7 are CUMULATIVE DELTAS, not
    standalone decks: a card's copies in a map's deck = the RUNNING SUM of its column values from Map 1
    through that map (clamped >=0). So decks evolve - e.g. Status: Normal starts at 6 and -1 each map;
    Double Event and new events accumulate. Map 6 & 7 share the last column. Each spreadsheet ROW is its
    own card (some names repeat with different descs, e.g. per-lane Reinforcements) so we sum per row."""
    ws = wbv['Campaign Event Cards']
    colspec = []                                   # (column, [positions it feeds]) in Map order
    for c in range(4, 10):
        h = ws.cell(1, c).value
        if isinstance(h, str) and 'Map' in h:
            colspec.append((c, [int(p) for p in re.findall(r'\d+', h)]))
    decks = {}
    for r in range(2, ws.max_row + 1):
        nm, desc, eff = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if not isinstance(nm, str) or not nm.strip():
            continue
        good = isinstance(eff, str) and eff.strip().lower() == 'good'
        dtext = desc.strip() if isinstance(desc, str) else ''
        running = 0
        for (c, positions) in colspec:             # walk Map 1..6&7 left-to-right, accumulating deltas
            cnt = ws.cell(r, c).value
            if isinstance(cnt, (int, float)):
                running += int(cnt)
            running = max(0, running)
            if running > 0:                        # this card is in each of this column's map decks
                for p in positions:
                    decks.setdefault(p, []).append({'name': nm.strip(), 'good': good, 'n': running, 'desc': dtext})
    return decks

def cell_kind(ws, r, c):
    col = ws.cell(r, c).fill.fgColor
    argb = col.rgb if (col is not None and col.type == 'rgb') else None
    return COLOR_KIND.get(argb)  # (kind, hazard) or None

def extract():
    wbv = openpyxl.load_workbook(XLSX, data_only=True)   # values
    wbs = openpyxl.load_workbook(XLSX, data_only=False)  # fills
    wv, wf = wbv['Campaign Maps'], wbs['Campaign Maps']

    # campaign decks (per-map enemy / treasure / event pools + the shared gather deck)
    enemy_decks = parse_enemy_decks(wbv)
    cull_rules = parse_cull_rules(wbv)
    treasure_pools = parse_treasures(wv)
    gather_deck = parse_gather_deck(wbv)
    event_decks = parse_event_decks(wbv)

    # locate scenario headers and board labels
    scen_rows = []       # (row, name, timed)
    boards = []          # (label, label_row, label_col, players)
    for r in range(1, wv.max_row + 1):
        for c in range(1, wv.max_column + 1):
            v = wv.cell(r, c).value
            if not isinstance(v, str):
                continue
            if 'Adventure Scenario' in v:
                nm = re.sub(r'^Adventure Scenario \d+:\s*', '', v).strip()
                timed = 'TIME' in v.upper()
                nm = re.sub(r'\s*\(TIME[^)]*\)', '', nm, flags=re.I).strip()
                scen_rows.append((r, nm, timed))
            elif re.match(r'^ADV\d', v.strip()):
                boards.append((v.strip(), r, c, wv.cell(r, c - 1).value))

    scen_rows.sort()
    def scenario_for(row):
        pick = None
        for (sr, nm, timed) in scen_rows:
            if sr <= row:
                pick = (nm, timed)
        return pick or ('Adventure', False)

    scenarios = {}  # name -> {name,timed,boards:[]}
    for (label, lr, lc, players) in boards:
        sname, timed = scenario_for(lr)
        cols = [lc - 1 + i for i in range(5)]            # 5 lane columns
        grid_rows = [lr + 1 + 2 * i for i in range(7)]   # 7 space rows, top(far)..bottom(home)
        kit_row = lr + 15

        lanes, structures, enemies, unmapped = [], [], [], []
        for li, col in enumerate(cols):
            spaces = []
            # bottom-up: idx 0 = home (bottom grid row) .. idx 6 = far/treasure (top)
            for idx, grow in enumerate(reversed(grid_rows)):
                ck = cell_kind(wf, grow, col)
                kind, hazard = ck if ck else ('plain', None)
                sp = {'kind': kind}
                if hazard:
                    sp['hazard'] = hazard
                spaces.append(sp)
                # overlay text on this cell
                tv = wv.cell(grow, col).value
                if isinstance(tv, str) and tv.strip():
                    t = tv.strip().upper()
                    if t in STRUCT:
                        structures.append({'lane': li, 'idx': idx, 'structId': STRUCT[t]})
                    elif t in ('NO TREASURE HERE',):
                        sp['kind'] = 'plain'  # a treasure-coloured space with no treasure
                    else:
                        eid = enemy_id(t)
                        if eid:
                            enemies.append({'lane': li, 'idx': idx, 'enemyDefId': eid})
                        else:
                            unmapped.append({'lane': li, 'idx': idx, 'text': tv.strip()})
            lanes.append({'name': f'Lane {li + 1}', 'spaces': spaces})  # match boards.json lane shape

        # pikmin kit (colour words on the kit row, across the lane columns)
        kit = []
        for col in cols:
            kv = wv.cell(kit_row, col).value
            if isinstance(kv, str) and kv.strip().upper() in PIKMIN:
                k = kv.strip().lower()
                if k not in kit:
                    kit.append(k)
        # basicColors = the board's starting colours (the whole kit). EVERY pikmin colour is
        # pellet-redeemable (pellets.json has red/blue/yellow/purple/white/rock/winged/ice), and
        # basicColors drives the pellet die AND the queen-candypop / colour-changing-posy targets.
        basics = list(kit) or ['red', 'blue', 'yellow']

        # the bracket note next to the label (looks like "(1,2)") is NOT players - it's which
        # STRUCTURES are BUILDABLE on this map: 1=bridge, 2=climbing stick, 3=tunnel. Walls are
        # never buildable. All three are hazard type "bridge" (a crossing), so they go in `bridges`.
        _sbi = {'1': 'bridge', '2': 'climbingstick', '3': 'tunnel'}
        build_structs = []
        for ch in str(players or ''):
            if ch in _sbi and _sbi[ch] not in build_structs:
                build_structs.append(_sbi[ch])

        bid = 'adv' + label.replace('ADV', '').replace('.', '_')
        mapkey = label.replace('ADV', 'A')                       # 'ADV1.1' -> 'A1.1'
        scenarios.setdefault(sname, {'name': sname, 'timed': timed, 'boards': []})
        pos = len(scenarios[sname]['boards']) + 1                # 1-based board position in its scenario

        # the map's treasure POOL (unplaced). scenario_adventure assigns these to the board's
        # treasure spaces RANDOMLY per run, leaving any surplus treasure space empty (still a
        # treasure space). A map may list fewer treasures than it has treasure spaces.
        treasure_pool = [{'id': t['id'], 'weight': t['weight'], 'name': t['name']}
                         for t in treasure_pools.get(mapkey, []) if t['id']]

        board = {
            'id': bid, 'label': label, 'players': str(players or ''), 'buildStructs': build_structs,
            'homeAnchored': True, 'basicColors': basics, 'kit': kit or ['red', 'blue', 'yellow'],
            'lanes': lanes,
            # placedStructures / placedEnemies = fixtures ALREADY on the board (distinct from a board
            # def's `structures` = the BUILDABLE set the game adds separately). enemies on a treasure
            # space are bosses; on an enemy space, regular enemies.
            'placedStructures': structures, 'placedEnemies': enemies, 'unmapped': unmapped,
            # campaign content: the map's enemy deck (fills bare enemy spaces + respawns), its treasures
            # (placed on the treasure spaces), the shared gather deck, and the random-event deck for this
            # board POSITION in the scenario (Map 6 & 7 share a deck). Consumed by scenario_adventure.
            'enemyDeck': enemy_decks.get(mapkey, []),
            # cullBefore = the ENEMY CULL applied to the carried-over campaign deck right before this
            # map is played (null on the first map / any map with no rule). {remove,thr,type}. The
            # player removes <remove> deck cards whose <type='hp'|'dmg'> stat is </= <thr>, then this
            # map's enemyDeck additions are mixed in. Persisted per-checkpoint in the adventure log.
            'cullBefore': cull_rules.get(mapkey),
            'treasures': treasure_pool,
            'gatherDeck': gather_deck,
            'eventDeck': event_decks.get(pos, event_decks.get(min(pos, 7), [])),
        }
        scenarios[sname]['boards'].append(board)

    return {'_comment': 'Adventure boards extracted from PikminCardReference.xlsx by tools/extract_adventure.py. '
                        'Home-anchored: treasure at the far end (idx 6), home near (idx 0). Per board: lanes, '
                        'placedStructures/placedEnemies (bosses on treasure spaces), treasures[] (from Campaign '
                        'Maps Z-AC, placed on treasure spaces), enemyDeck[] (Campaign N Enemies), gatherDeck[] '
                        '(shared 12 campaign cards), eventDeck[] (Campaign Event Cards, by board position; drawn '
                        'one/turn). unmapped[] = overlay text with no game id yet.',
            'scenarios': list(scenarios.values())}

if __name__ == '__main__':
    data = extract()
    with open(OUT, 'w') as f:
        json.dump(data, f, indent=1)
    ns = len(data['scenarios'])
    nb = sum(len(s['boards']) for s in data['scenarios'])
    print(f'wrote {OUT}: {ns} scenarios, {nb} boards')
    for s in data['scenarios']:
        print(f"  {s['name']!r} (timed={s['timed']}): {[b['label'] for b in s['boards']]}")
